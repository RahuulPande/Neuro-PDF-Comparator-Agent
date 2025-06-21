"""
HTMLReportGenerator: Generates visually appealing HTML reports for PDF comparison results.
"""

from typing import Dict, Any
import datetime
import tempfile
import os

class HTMLReportGenerator:
    """Generates HTML reports with modern CSS and interactive elements."""
    def generate_report(self, differences: Dict[str, Any], intelligent_summary: Dict[str, Any], config: Dict[str, Any]) -> str:
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        css = self._modern_css(config)
        
        # Extract LLM analysis results
        summary = intelligent_summary.get("intelligent_summary", "")
        severity_assessment = intelligent_summary.get("severity_assessment", {})
        pattern_analysis = intelligent_summary.get("pattern_analysis", {})
        
        html = f"""
        <html>
        <head>
            <title>PDF Comparison Report</title>
            <style>{css}</style>
        </head>
        <body>
            <div class='header'>
                <h1>PDF Comparison Report</h1>
                <div class='meta'>Generated: {now}</div>
            </div>
            
            <div class='summary'>
                <h2>AI Analysis Summary</h2>
                <p>{summary}</p>
            </div>
            
            {self._render_severity_assessment(severity_assessment)}
            {self._render_pattern_analysis(pattern_analysis)}
            
            <div class='results'>
                <h2>Detailed Differences</h2>
                {self._render_differences(differences)}
            </div>
        </body>
        </html>
        """
        return html

    def generate_excel_report(self, differences: Dict[str, Any]) -> str:
        """Generate a structured Excel report using openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            
            # Remove default sheet and create our own
            if wb.active:
                wb.remove(wb.active)
            
            # Create main differences worksheet
            ws_diff = wb.create_sheet("Differences")
            headers = ["File", "Category", "Details"]
            ws_diff.append(headers)
            for cell in ws_diff[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            for filename, diff_categories in differences.items():
                if "error" in diff_categories:
                    ws_diff.append([filename, "Error", diff_categories["error"]])
                    continue

                for category, diff_list in diff_categories.items():
                    for detail in diff_list:
                        ws_diff.append([filename, category.capitalize(), detail])
            
            # Auto-adjust column widths for differences sheet
            for col_idx, column_cells in enumerate(ws_diff.columns, 1):
                from openpyxl.cell import MergedCell
                max_length = 0
                column_letter = None
                for cell in column_cells:
                    if isinstance(cell, MergedCell):
                        continue # Skip merged cells
                    if cell.value:
                        if not column_letter:
                           column_letter = cell.column_letter
                        max_length = max(max_length, len(str(cell.value)))

                if column_letter and max_length > 0:
                    adjusted_width = (max_length + 2)
                    ws_diff.column_dimensions[column_letter].width = adjusted_width

            # Create summary worksheet
            ws_summary = wb.create_sheet("Summary")
            ws_summary.append(["Metric", "Value"])
            ws_summary.append(["Total Files", len(differences)])
            
            total_changes = 0
            for diff_categories in differences.values():
                if isinstance(diff_categories, dict):
                    for category, diff_list in diff_categories.items():
                        if category != "error":
                            total_changes += len(diff_list)
            
            ws_summary.append(["Total Changes", total_changes])
            
            # Style summary sheet
            for cell in ws_summary[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
            os.close(fd)
            wb.save(temp_path)
            return temp_path
            
        except ImportError:
            return self._generate_csv_fallback(differences)
        except Exception as e:
            return self._generate_csv_fallback(differences)

    def _generate_csv_fallback(self, differences: Dict[str, Any]) -> str:
        """Generate CSV report as fallback."""
        import tempfile
        import os
        
        # Create a simple CSV-like report
        report_lines = ["Filename,Change Type,Details"]
        for filename, diff in differences.items():
            if isinstance(diff, dict):
                for change_type, changes in diff.items():
                    if isinstance(changes, list):
                        for change in changes:
                            report_lines.append(f"{filename},{change_type},{change}")
                    else:
                        report_lines.append(f"{filename},{change_type},{changes}")
        
        # Save to temporary file
        fd, temp_path = tempfile.mkstemp(suffix='.csv')
        with os.fdopen(fd, 'w') as f:
            f.write('\n'.join(report_lines))
        
        return temp_path

    def _render_differences(self, differences: Dict[str, Any]) -> str:
        """Renders the structured differences into a readable HTML format."""
        html = ""
        for fname, diff_categories in differences.items():
            html += f"<div class='file-section'><button class='collapsible'>{fname}</button><div class='content'>"
            
            if "error" in diff_categories:
                html += f"<div class='diff-remove'>Error: {diff_categories['error']}</div>"
            else:
                html += "<ul class='diff-list'>"
                # Process text differences first
                if "text" in diff_categories and diff_categories["text"]:
                    for item in diff_categories["text"]:
                        html += f"<li>{item}</li>"
                # Process other categories
                for category, diff_list in diff_categories.items():
                    if category != "text":
                        for item in diff_list:
                             html += f"<li>{item}</li>"
                html += "</ul>"

            html += "</div></div>"
        
        # Add collapsible JS
        html += """
        <script>
        var coll = document.getElementsByClassName('collapsible');
        for (var i = 0; i < coll.length; i++) {
            coll[i].addEventListener('click', function() {
                this.classList.toggle('active');
                var content = this.nextElementSibling;
                if (content.style.display === 'block') {
                    content.style.display = 'none';
                } else {
                    content.style.display = 'block';
                }
            });
        }
        </script>
        """
        return html

    def _render_severity_assessment(self, severity_assessment: Dict[str, Any]) -> str:
        """Render severity assessment in HTML format."""
        if not severity_assessment:
            return ""
        
        html = '<div class="severity-assessment"><h2>Severity Assessment</h2>'
        
        for severity_level in ["Critical", "Major", "Minor"]:
            files = severity_assessment.get(severity_level, [])
            if files:
                color_class = {
                    "Critical": "severity-critical",
                    "Major": "severity-major", 
                    "Minor": "severity-minor"
                }.get(severity_level, "")
                
                html += f'<div class="severity-level {color_class}">'
                html += f'<h3>{severity_level} Changes</h3>'
                html += '<ul>'
                for file in files:
                    html += f'<li>{file}</li>'
                html += '</ul></div>'
        
        html += '</div>'
        return html
    
    def _render_pattern_analysis(self, pattern_analysis: Dict[str, Any]) -> str:
        """Render pattern analysis in HTML format."""
        if not pattern_analysis:
            return ""
        
        html = '<div class="pattern-analysis"><h2>Pattern Analysis</h2>'
        
        # Render patterns
        patterns = pattern_analysis.get("patterns", [])
        if patterns:
            html += '<div class="analysis-section"><h3>Identified Patterns</h3><ul>'
            for pattern in patterns:
                html += f'<li>{pattern}</li>'
            html += '</ul></div>'
        
        # Render trends
        trends = pattern_analysis.get("trends", [])
        if trends:
            html += '<div class="analysis-section"><h3>Observed Trends</h3><ul>'
            for trend in trends:
                html += f'<li>{trend}</li>'
            html += '</ul></div>'
        
        # Render insights
        insights = pattern_analysis.get("insights", [])
        if insights:
            html += '<div class="analysis-section"><h3>Business Insights</h3><ul>'
            for insight in insights:
                html += f'<li>{insight}</li>'
            html += '</ul></div>'
        
        html += '</div>'
        return html

    def _modern_css(self, config: Dict[str, Any]) -> str:
        return """
        body { font-family: 'Segoe UI', Arial, sans-serif; background: #f8f9fa; color: #212529; }
        .header { background: linear-gradient(90deg, #1f77b4, #ff7f0e); color: white; padding: 1.5em; box-shadow: 0 2px 8px #0002; }
        .meta { font-size: 0.9em; opacity: 0.8; }
        .summary { background: #fff; margin: 2em 0; padding: 1.5em; border-radius: 8px; box-shadow: 0 1px 4px #0001; }
        .severity-assessment { background: #fff; margin: 2em 0; padding: 1.5em; border-radius: 8px; box-shadow: 0 1px 4px #0001; }
        .pattern-analysis { background: #fff; margin: 2em 0; padding: 1.5em; border-radius: 8px; box-shadow: 0 1px 4px #0001; }
        .severity-level { margin: 1em 0; padding: 1em; border-radius: 6px; border-left: 5px solid; }
        .severity-critical { background: #ffe6e6; border-left-color: #dc3545; }
        .severity-major { background: #fff3cd; border-left-color: #ffc107; }
        .severity-minor { background: #d1ecf1; border-left-color: #17a2b8; }
        .analysis-section { margin: 1em 0; }
        .analysis-section h3 { color: #495057; margin-bottom: 0.5em; }
        .analysis-section ul { margin: 0; padding-left: 1.5em; }
        .analysis-section li { margin: 0.5em 0; }
        .results { background: #fff; padding: 1.5em; border-radius: 8px; box-shadow: 0 1px 4px #0001; }
        h1, h2 { margin-top: 0; }
        .diff-list { list-style-type: none; padding-left: 0; }
        .diff-list li { background: #f1f1f1; margin-bottom: 5px; padding: 10px; border-radius: 4px; border-left: 5px solid #ff7f0e; }
        .diff-remove { color: #dc3545; }
        .file-section { margin-bottom: 2em; }
        .collapsible { width: 100%; text-align: left; cursor: pointer; padding: 1em; border: none; background: #e9ecef; border-radius: 4px; margin-bottom: 0.5em; font-size: 1.1em; font-weight: bold; }
        .content { display: none; padding: 0.5em 1em; border: 1px solid #e9ecef; border-top: none; border-radius: 0 0 4px 4px;}
        .active, .collapsible:hover { background: #dee2e6; }
        """ 